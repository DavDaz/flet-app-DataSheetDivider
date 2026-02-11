# Combined_script.py
import os
import pandas as pd
import re
import datetime
from copy import copy
from collections import defaultdict
from dotenv import load_dotenv
import flet as ft
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter, range_boundaries

# Cargar variables de entorno desde el archivo .env
load_dotenv()

def _copy_cell(src_cell, ws_dst, dst_row):
    """Copia una celda con su valor y estilo completo a la fila dst_row de ws_dst."""
    dst_cell = ws_dst.cell(row=dst_row, column=src_cell.column, value=src_cell.value)
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)


def _add_tables(ws_src, ws_dst, new_row_count, display_name_prefix="T"):
    """
    Recrea las tablas de ws_src en ws_dst con el rango ajustado a new_row_count filas.
    display_name_prefix debe ser único dentro del workbook destino.
    """
    for i, src_table in enumerate(ws_src.tables.values()):
        min_col, _min_row, max_col, _max_row = range_boundaries(src_table.ref)
        new_ref = f"{get_column_letter(min_col)}1:{get_column_letter(max_col)}{new_row_count}"

        safe_prefix = re.sub(r'[^A-Za-z0-9_]', '_', display_name_prefix)[:28]
        display_name = f"{safe_prefix}_{i}" if i > 0 else safe_prefix

        new_table = Table(displayName=display_name, ref=new_ref)
        if src_table.tableStyleInfo:
            new_table.tableStyleInfo = TableStyleInfo(
                name=src_table.tableStyleInfo.name,
                showFirstColumn=src_table.tableStyleInfo.showFirstColumn,
                showLastColumn=src_table.tableStyleInfo.showLastColumn,
                showRowStripes=src_table.tableStyleInfo.showRowStripes,
                showColumnStripes=src_table.tableStyleInfo.showColumnStripes,
            )
        ws_dst.add_table(new_table)


def _copy_sheet(ws_src, ws_dst):
    """Copia todas las celdas, anchos de columna, altos de fila y celdas combinadas de ws_src a ws_dst."""
    for row in ws_src.iter_rows():
        for cell in row:
            _copy_cell(cell, ws_dst, dst_row=cell.row)

    for col_letter, col_dim in ws_src.column_dimensions.items():
        ws_dst.column_dimensions[col_letter].width = col_dim.width

    for row_idx, row_dim in ws_src.row_dimensions.items():
        ws_dst.row_dimensions[row_idx].height = row_dim.height

    for merged_range in ws_src.merged_cells.ranges:
        ws_dst.merge_cells(str(merged_range))

    # Copiar tablas con el mismo rango (copia completa de la hoja)
    total_rows = ws_src.max_row
    _add_tables(ws_src, ws_dst, new_row_count=total_rows, display_name_prefix=ws_dst.title)


def sanitize_sheet_name(name):
    """
    Sanitiza el nombre de una hoja de Excel para evitar caracteres no válidos.
    Excel no permite: [ ] * ? : / \
    También limita la longitud a 31 caracteres.
    """
    if name is None:
        return "Sin_nombre"

    # Convertir a string si no lo es
    name = str(name)

    # Reemplazar caracteres no válidos con guiones bajos
    invalid_chars = r'[\[\]\*\?:/\\]'
    sanitized = re.sub(invalid_chars, '_', name)

    # Eliminar caracteres no imprimibles
    sanitized = ''.join(c for c in sanitized if c.isprintable())

    # Si está vacío después de sanitizar, usar un nombre predeterminado
    if not sanitized or sanitized.isspace():
        sanitized = "Hoja_Sin_Nombre"

    # Limitar a 31 caracteres (límite de Excel)
    return sanitized[:31]


def exportar_ventanas_xlsx(file_path, carpeta_principal):
    """
    Exporta cada hoja del archivo Excel a un archivo individual
    dentro de una subcarpeta de la carpeta principal proporcionada.
    
    Args:
        file_path: Ruta al archivo Excel que contiene las hojas a exportar
        carpeta_principal: Carpeta principal donde se creará la subcarpeta para los archivos individuales
    
    Returns:
        Ruta a la subcarpeta donde se guardaron los archivos individuales
    """
    # Cargar el archivo xlsx preservando estilos
    wb_combined = openpyxl.load_workbook(file_path, data_only=True)

    # Obtener el nombre del archivo sin la extensión
    file_name = os.path.splitext(os.path.basename(file_path))[0]

    # Crear una subcarpeta para almacenar los archivos individuales dentro de la carpeta principal
    folder_name = f"Separados-{file_name}"
    full_folder_path = os.path.join(carpeta_principal, folder_name)
    os.makedirs(full_folder_path, exist_ok=True)

    # Iterar sobre todas las hojas en el archivo xlsx
    for sheet_name in wb_combined.sheetnames:
        try:
            # Generar el nombre de archivo exportado
            export_file_name = f"{file_name}-{sheet_name}.xlsx"
            export_file_path = os.path.join(full_folder_path, export_file_name)

            # Crear un workbook individual y copiar la hoja con formato
            wb_individual = openpyxl.Workbook()
            wb_individual.remove(wb_individual.active)
            ws_dst = wb_individual.create_sheet(title=sheet_name)
            _copy_sheet(wb_combined[sheet_name], ws_dst)
            wb_individual.save(export_file_path)
        except Exception as e:
            print(f"Error al exportar la hoja {sheet_name}: {str(e)}")
            continue

    print(f"Se han exportado las ventanas por xlsx separados en la carpeta: {full_folder_path}")
    return full_folder_path

def get_default_folder_name():
    """
    Genera un nombre de carpeta por defecto con la fecha y hora actual.
    Formato: Resultados_YYYY-MM-DD_HH-MM-SS
    """
    now = datetime.datetime.now()
    return f"Resultados_{now.strftime('%Y-%m-%d_%H-%M-%S')}"

def main(page):
    # Contenedor para mostrar resultados
    resultados_container = ft.Container(
        content=None,
        padding=10,
        visible=False
    )
    
    # Indicador de carga para las hojas
    hojas_cargando = ft.Container(
        content=ft.Row([
            ft.ProgressRing(width=16, height=16, stroke_width=2),
            ft.Text("Cargando hojas del Excel...", size=12)
        ]),
        visible=False,
        padding=ft.padding.only(left=10, top=5)
    )
    
    # Indicador de carga para las columnas
    columnas_cargando = ft.Container(
        content=ft.Row([
            ft.ProgressRing(width=16, height=16, stroke_width=2),
            ft.Text("Cargando columnas...", size=12)
        ]),
        visible=False,
        padding=ft.padding.only(left=10, top=5)
    )
    
    # Función para limpiar los campos y preparar para procesar otro archivo
    def limpiar_campos(e=None):
        input_excel.value = ""
        sheets_dropdown.options.clear()
        sheets_dropdown.disabled = True
        columns_dropdown.options.clear()
        columns_dropdown.disabled = True
        input_carpet.value = ""
        resultados_container.visible = False
        columnas_cargando.visible = False
        hojas_cargando.visible = False
        page.update()
    
    # Variables para almacenar las hojas y columnas
    sheets_dropdown = ft.Dropdown(
        label="Seleccionar Hoja",
        width=400,
        disabled=True,
    )
    
    columns_dropdown = ft.Dropdown(
        label="Seleccionar Columna",
        width=400,
        disabled=True,
    )
    
    # Crear un contenedor para el campo de texto y el botón de selección
    input_excel = ft.TextField(
        label="Nombre del Archivo Excel", 
        read_only=True,
        expand=True
    )
    
    def cargar_hojas_excel(file_path):
        try:
            # Mostrar indicador de carga
            hojas_cargando.visible = True
            sheets_dropdown.disabled = True
            page.update()
            
            # Limpiar dropdowns anteriores
            sheets_dropdown.options.clear()
            columns_dropdown.options.clear()
            columns_dropdown.disabled = True
            
            # Cargar el archivo Excel
            xls = pd.ExcelFile(file_path)
            
            # Agregar las hojas al dropdown
            for sheet in xls.sheet_names:
                sheets_dropdown.options.append(ft.dropdown.Option(sheet))
            
            # Ocultar indicador de carga y habilitar el dropdown
            hojas_cargando.visible = False
            sheets_dropdown.disabled = False
            sheets_dropdown.value = None
            page.update()
        except Exception as e:
            # Ocultar indicador de carga en caso de error
            hojas_cargando.visible = False
            page.snack_bar = ft.SnackBar(
                content=ft.Text(f"Error al cargar el archivo Excel: {str(e)}"),
                bgcolor=ft.colors.RED_400
            )
            page.snack_bar.open = True
            page.update()
    
    def on_sheet_change(e):
        try:
            if not input_excel.value or not sheets_dropdown.value:
                return
            
            # Mostrar indicador de carga
            columnas_cargando.visible = True
            columns_dropdown.disabled = True
            page.update()
                
            # Limpiar dropdown de columnas
            columns_dropdown.options.clear()
            
            # Cargar las columnas de la hoja seleccionada
            df = pd.read_excel(input_excel.value, sheet_name=sheets_dropdown.value)
            
            # Agregar las columnas al dropdown
            for column in df.columns:
                columns_dropdown.options.append(ft.dropdown.Option(column))
            
            # Ocultar indicador de carga y habilitar el dropdown
            columnas_cargando.visible = False
            columns_dropdown.disabled = False
            page.update()
        except Exception as e:
            # Ocultar indicador de carga en caso de error
            columnas_cargando.visible = False
            page.snack_bar = ft.SnackBar(
                content=ft.Text(f"Error al cargar las columnas: {str(e)}"),
                bgcolor=ft.colors.RED_400
            )
            page.snack_bar.open = True
            page.update()
    
    # Asignar el evento de cambio al dropdown de hojas
    sheets_dropdown.on_change = on_sheet_change
    
    def on_file_picker_result(e: ft.FilePickerResultEvent):
        if e.files:
            # Actualizar el campo de texto con la ruta del archivo seleccionado
            input_excel.value = e.files[0].path
            
            # Cargar las hojas del Excel
            cargar_hojas_excel(input_excel.value)
            
            page.update()
    
    # Crear el FilePicker
    file_picker = ft.FilePicker(on_result=on_file_picker_result)
    page.overlay.append(file_picker)
    
    def select_excel_file(e):
        file_picker.pick_files(
            dialog_title="Seleccionar archivo Excel",
            allowed_extensions=["xlsx", "xls"],
            file_type=ft.FilePickerFileType.CUSTOM
        )
    
    def mostrar_progreso(mensaje):
        resultados_container.content = ft.Column([
            ft.ProgressBar(width=400),
            ft.Text(mensaje)
        ])
        resultados_container.visible = True
        page.update()
    
    def mostrar_resultado(titulo, mensajes, es_error=False):
        contenido = [ft.Text(titulo, size=20, weight=ft.FontWeight.BOLD, color=ft.colors.RED if es_error else None)]
        
        for mensaje in mensajes:
            contenido.append(ft.Text(mensaje, style=ft.TextThemeStyle.BODY_MEDIUM))
        
        contenido.append(ft.ElevatedButton("Procesar otro archivo", on_click=limpiar_campos))
        
        resultados_container.content = ft.Column(contenido)
        resultados_container.visible = True
        page.update()
    
    def btn_click(e):
        if not input_excel.value or not sheets_dropdown.value or not columns_dropdown.value:
            # Mostrar mensajes de error en los campos vacíos
            if not input_excel.value:
                page.snack_bar = ft.SnackBar(content=ft.Text("Por favor selecciona un archivo Excel"))
                page.snack_bar.open = True
            elif not sheets_dropdown.value:
                page.snack_bar = ft.SnackBar(content=ft.Text("Por favor selecciona una hoja"))
                page.snack_bar.open = True
            elif not columns_dropdown.value:
                page.snack_bar = ft.SnackBar(content=ft.Text("Por favor selecciona una columna"))
                page.snack_bar.open = True
            page.update()
        else:
            archivo_excel = input_excel.value
            nombre_sede_column = columns_dropdown.value
            
            # Si no se proporciona un nombre de carpeta, generar uno por defecto
            carpeta_resultado = input_carpet.value.strip() if input_carpet.value and input_carpet.value.strip() else get_default_folder_name()

            # Verificar si el archivo Excel existe
            if not os.path.isfile(archivo_excel):
                mostrar_resultado(
                    "Error al procesar el archivo", 
                    ["El archivo Excel especificado no se encuentra. Por favor, verifique la ruta y vuelva a intentarlo."],
                    es_error=True
                )
            else:
                try:
                    # Cargar la hoja seleccionada con openpyxl para preservar estilos
                    wb_src = openpyxl.load_workbook(archivo_excel, data_only=True)
                    ws_src = wb_src[sheets_dropdown.value]

                    # Obtener la fila de encabezado e identificar el índice de la columna elegida
                    header_row = list(ws_src.iter_rows(min_row=1, max_row=1))[0]
                    col_idx = next((c.column for c in header_row if str(c.value) == nombre_sede_column), None)

                    if col_idx is None:
                        mostrar_resultado(
                            "Error al procesar el archivo",
                            ["El nombre de la columna especificado no existe en el archivo Excel. Por favor, verifique el nombre de la columna y vuelva a intentarlo."],
                            es_error=True
                        )
                    else:
                        # Agrupar filas de datos por valor único de la columna (manteniendo orden de aparición)
                        groups = defaultdict(list)
                        seen_values = []
                        for row in ws_src.iter_rows(min_row=2):
                            val = row[col_idx - 1].value
                            if val not in groups:
                                seen_values.append(val)
                            groups[val].append(row)

                        # Obtener el directorio donde se encuentra el archivo Excel original
                        directorio_excel = os.path.dirname(archivo_excel)

                        # Crear la carpeta principal de resultados en la misma ubicación que el archivo Excel
                        ruta_carpeta_principal = os.path.join(directorio_excel, carpeta_resultado)
                        if not os.path.exists(ruta_carpeta_principal):
                            os.makedirs(ruta_carpeta_principal)

                        # Mostrar mensaje de progreso
                        mostrar_progreso("Procesando archivos...")

                        # Crear un nuevo archivo Excel con openpyxl preservando el formato original
                        nombre_archivo_resultado = f"{os.path.basename(archivo_excel)}"
                        ruta_resultado = os.path.join(ruta_carpeta_principal, nombre_archivo_resultado)

                        # Diccionario para llevar registro de nombres de hojas ya utilizados
                        used_sheet_names = {}

                        wb_out = openpyxl.Workbook()
                        wb_out.remove(wb_out.active)

                        for corregimiento in seen_values:
                            try:
                                # Sanitizar el nombre de la hoja
                                sheet_name = sanitize_sheet_name(corregimiento)

                                # Manejar duplicados agregando un número
                                if sheet_name in used_sheet_names:
                                    used_sheet_names[sheet_name] += 1
                                    sheet_name = f"{sheet_name}_{used_sheet_names[sheet_name]}"
                                else:
                                    used_sheet_names[sheet_name] = 0

                                ws_out = wb_out.create_sheet(title=sheet_name)

                                # Copiar fila de encabezado con formato
                                for cell in header_row:
                                    _copy_cell(cell, ws_out, dst_row=1)

                                # Copiar filas de datos del grupo con formato
                                for dst_row, src_row in enumerate(groups[corregimiento], start=2):
                                    for cell in src_row:
                                        _copy_cell(cell, ws_out, dst_row=dst_row)

                                # Copiar anchos de columna del original
                                for col_letter, col_dim in ws_src.column_dimensions.items():
                                    ws_out.column_dimensions[col_letter].width = col_dim.width

                                # Recrear tablas del original con rango ajustado a las filas del grupo
                                new_row_count = 1 + len(groups[corregimiento])  # encabezado + filas del grupo
                                _add_tables(ws_src, ws_out, new_row_count=new_row_count, display_name_prefix=sheet_name)

                            except Exception as e:
                                print(f"Error al procesar el valor '{corregimiento}': {str(e)}")
                                continue

                        wb_out.save(ruta_resultado)

                        # Llamar a la función para exportar las ventanas y obtener la ruta de la carpeta de resultados
                        # Pasar la carpeta principal como parámetro para que los archivos individuales se creen dentro de ella
                        carpeta_archivos = exportar_ventanas_xlsx(ruta_resultado, ruta_carpeta_principal)

                        # Mostrar mensaje de completado
                        mostrar_resultado(
                            "¡Proceso completado!",
                            [
                                f"Las ventanas se han exportado por separado en archivos Excel en:",
                                f"Carpeta principal: {ruta_carpeta_principal}",
                                f"1. Archivo combinado: {nombre_archivo_resultado}",
                                f"2. Archivos individuales: {os.path.basename(carpeta_archivos)}"
                            ]
                        )
                except Exception as e:
                    # Manejar cualquier error inesperado
                    mostrar_resultado(
                        "Error al procesar el archivo",
                        [f"Ocurrió un error: {str(e)}"],
                        es_error=True
                    )
    
    excel_file_row = ft.Row(
        controls=[
            input_excel,
            ft.IconButton(
                icon=ft.icons.FOLDER_OPEN,
                tooltip="Seleccionar archivo Excel",
                on_click=select_excel_file
            )
        ]
    )
    
    input_carpet = ft.TextField(
        label="Escribe un nuevo Nombre para crear la carpeta de resultados (opcional)",
        hint_text="Dejar en blanco para generar automáticamente"
    )

    page.add(
        ft.Text("Divisor de Hojas de Excel", size=20, weight=ft.FontWeight.BOLD),
        excel_file_row,
        hojas_cargando,
        sheets_dropdown,
        columnas_cargando,
        columns_dropdown,
        input_carpet, 
        ft.ElevatedButton("Ejecutar Proceso!", on_click=btn_click),
        resultados_container
    )

ft.app(target=main)
